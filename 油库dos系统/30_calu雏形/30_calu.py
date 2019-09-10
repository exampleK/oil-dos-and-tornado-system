import math,time,sys
import xlrd #read execl
from openpyxl import Workbook #write execl
from openpyxl import load_workbook
oil_R = 1.2
# oil_H = 2.22 ---->L=29.998068441092098
oil_L = 6.05
oil_C = 0.6
# oli_current = 31
global update_oil1,update_oil2,execl_oil
update_oil1 = 0
update_oil2 = 0

L = 0
def intrduce():
    update_oil1 = read_oil_log_execl()[0]
    update_oil2 = read_oil_log_execl()[1]
    menu_1 = ("""                 welcome 30^ oil calculate system
                                                                author:wa0gka1
    ---Panel---""")
    menu_2 = "current:"+str(update_oil1+update_oil2)
    menu_3 = "---oil_H1:"+str(update_oil1)
    menu_4 = "---oil_H2:"+str(update_oil2)
    menu_5 = "---if current-7000:"+str(update_oil1+update_oil2-0.7)
    menu_5_1 = "---if execl-7000:"+str((execl_11()-7000)/1000)
    menu_6 = ("""---menu---
    '0' update menu current oil
    '1' calculate oil'L 
    '2' auto calculate remaining oil
    '3' quit()""")
    menu_list =[menu_1,menu_2,menu_3,menu_4,menu_5,menu_5_1,menu_6]
    for i in menu_list:
        print(i)
    if execl_oil < 10000:
        print("*"*30,"WARNING!!!,execl oil < 10000L")
def calu_oil_L(oil_H):
    L = (oil_L*((math.pi * oil_R**2)/2-(oil_R-oil_H)*math.sqrt(2*oil_H*oil_R-oil_H**2)-oil_R**2*math.asin(1-oil_H/oil_R))+(math.pi*oil_C)/(3*oil_R)*(3*oil_R**2*oil_H-oil_R**3+(oil_R-oil_H)**3))
    return L
def execl_11():
    a = (sys.path[0]+'/oil_churu.xls').replace("\\",'/')
    # 读取文件
    data = xlrd.open_workbook(sys.path[0]+'/oil_churu.xls')
    # 通过索引顺序获取 ：sheet 0 就是第一个
    table = data.sheets()[0]
    table.col_values(0)
    test_list1=[]
    for i in range(len(table.col_values(10))):
        if i > 2:
            test_list1.append(table.col_values(10)[i])
        else:
            continue
    # print((test_list1))
    return sum(test_list1)
def oil_H_add():
    wb = load_workbook(sys.path[0]+'\oil_log.xlsx')
    ws=wb.active# 激活
    # column=ws.max_column #获取行数
    # row=ws.max_row #获取列数
    row = ws.max_row
    ws.cell(row+1,1).value=time_loacl()
    ws.cell(row+1,2).value=update_oil1
    ws.cell(row+1,3).value=update_oil2
    ws.cell(row+1,4).value=update_oil1+update_oil2
    ws.cell(row+1,5).value=update_oil1+update_oil2-0.7
    # print(row)
    wb.save(r"C:\\oil_log.xlsx")
    print("execl save success!")
def read_oil_log_execl():
    # 读取文件
    data = xlrd.open_workbook(sys.path[0]+'\oil_log.xlsx')
    # 通过索引顺序获取 ：sheet 0 就是第一个
    table = data.sheets()[0] #第0个sheet
    list_oil_log = []
    table.col_values(3)[len(table.col_values(3))-1]# table.col_values(3)[len(table.col_values(3))-1]第3列的最后一行数值
    list_oil_log.append(table.col_values(1)[len(table.col_values(1))-1])
    list_oil_log.append(table.col_values(2)[len(table.col_values(2))-1])
    list_oil_log.append(table.col_values(3)[len(table.col_values(3))-1])
    return list_oil_log
def time_loacl():
    local_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    return local_time
# -----------------------------------main-------------------------------------------------
if __name__=='__main__':
    # oil_H_add()
    execl_oil=execl_11()
    intrduce()
    # execl_oil=execl_11()
    while 1:
        command = int(input("please Command:"))
        if command == 0:
            intrduce()
        elif command == 1:
            command_value = input("""are you update menu:
                Yes->enter
                No ->whatever keyword""")
            if command_value:
                print(bool(command_value))
                # oil_H = float(input("please input 1-2_oil's high(m):"))
                oil_H1 = float(input("please input 1_oil's high(m):"))
                oil_current1 = (calu_oil_L(oil_H1))
                oil_H2 = float(input("please input 2_oil's high(m):"))
                oil_current2 = (calu_oil_L(oil_H2))
                print("oil_H1:",oil_current1)
                print("oil_H2:",oil_current2)
                # print(oil_L*((math.pi * oil_R**2)/2-(oil_R-oil_H)*math.sqrt(2*oil_H*oil_R-oil_H**2)-oil_R**2*math.asin(1-oil_H/oil_R))+(math.pi*oil_C)/(3*oil_R)*(3*oil_R**2*oil_H-oil_R**3+(oil_R-oil_H)**3))
            else:
                print(bool(command_value))
                oil_H1 = float(input("please input 1_oil's high(m):"))
                oil_current1 = (calu_oil_L(oil_H1))
                oil_H2 = float(input("please input 2_oil's high(m):"))
                oil_current2 = (calu_oil_L(oil_H2))
                print("oil_H1:",oil_current1)
                print("oil_H2:",oil_current2)
                update_oil1 = oil_current1
                update_oil2 = oil_current2
                oil_H_add()
                print("update success!")
                intrduce()

        elif command == 2:
            pass
        elif command == 3:
            quit()
        else:
            print("---command error---")
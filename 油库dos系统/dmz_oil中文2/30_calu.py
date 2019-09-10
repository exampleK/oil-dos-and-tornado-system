import math,time,os
import xlrd #read execl
import sys
from openpyxl import Workbook #write execl
from openpyxl import load_workbook
oil_R = 1.2
# oil_H = 2.22 ---->L=29.998068441092098
oil_L = 6.05
oil_C = 0.6
# oli_current = 31
global update_oil1,update_oil2,execl_oil
update_oil1 = 0
update_oil2 = 0

L = 0
def local_dir():
    # a =  (sys.path[0]+"\oil_churu1.xls")
    # print(r"%s"%a)
    # a = (sys.path[0]+"\\")
    a = r"c:\\"
    return a

def intrduce():
    update_oil1 = read_oil_log_execl()[0]
    update_oil2 = read_oil_log_execl()[1]
    update_oil3 = read_oil_log_execl()[3]
    menu_1 = ("""                 欢迎进入 30立方油罐 液位 管理系统
                                                                作者:wa0g ka1
                                    ---菜单---""")
    menu_2_1 = "上一次 液位 测量 的时间:"+str(update_oil3)
    menu_2 = "上一次 液位 测量 的油量剩余:"+str(update_oil1+update_oil2)
    menu_3 = "当前1号油罐的剩余量:"+str(update_oil1)
    menu_4 = "当前2号油罐的剩余量:"+str(update_oil2)
    menu_5 = "根据上一次液位测量的剩余量-打不出来的油量（7000）="+str(update_oil1+update_oil2-0.7)
    menu_5_1 = "根据油量出入execl表的剩余总和剩余量-打不出来的油量（7000）="+str((execl_11()-7000)/1000)
    menu_6 = ("""                                    ---选项---
    '0' 更新execl表的剩余油量
    '1' 根据液位公式计算相对精确剩余油量
    # '2' auto calculate remaining oil（暂未开通）
    '3' 退出本系统""")
    menu_list =[menu_1,menu_2_1,menu_2,menu_3,menu_4,menu_5,menu_5_1,menu_6]
    for i in menu_list:
        print(i)
    if execl_oil < 10000:
        print("*"*30,"警告！！！ 油量少于10000L，请联系机电物资科")
def calu_oil_L(oil_H):
    L = (oil_L*((math.pi * oil_R**2)/2-(oil_R-oil_H)*math.sqrt(2*oil_H*oil_R-oil_H**2)-oil_R**2*math.asin(1-oil_H/oil_R))+(math.pi*oil_C)/(3*oil_R)*(3*oil_R**2*oil_H-oil_R**3+(oil_R-oil_H)**3))
    return L
def execl_11():
    # 读取文件
    data = xlrd.open_workbook(local_dir()+'油料出入.xls')
    # 通过索引顺序获取 ：sheet 0 就是第一个
    table = data.sheets()[0]
    table.col_values(0)
    test_list1=[]
    for i in range(len(table.col_values(10))):
        if i > 2:
            test_list1.append(table.col_values(10)[i])
        else:
            continue
    # print((test_list1))
    return sum(test_list1)
def oil_H_add():
    wb = load_workbook(local_dir()+'oil_log.xlsx')
    ws=wb.active# 激活
    # column=ws.max_column #获取行数
    # row=ws.max_row #获取列数
    row = ws.max_row
    ws.cell(row+1,1).value=time_loacl()
    ws.cell(row+1,2).value=update_oil1
    ws.cell(row+1,3).value=update_oil2
    ws.cell(row+1,4).value=update_oil1+update_oil2
    ws.cell(row+1,5).value=update_oil1+update_oil2-0.7
    # print(row)
    wb.save(local_dir()+"oil_log.xlsx")
    print("execl保存成功!")
def read_oil_log_execl():
    # 读取文件
    data = xlrd.open_workbook(local_dir()+'oil_log.xlsx')
    # 通过索引顺序获取 ：sheet 0 就是第一个
    table = data.sheets()[0] #第0个sheet
    list_oil_log = []
    table.col_values(3)[len(table.col_values(3))-1]# table.col_values(3)[len(table.col_values(3))-1]第3列的最后一行数值
    list_oil_log.append(table.col_values(1)[len(table.col_values(1))-1])
    list_oil_log.append(table.col_values(2)[len(table.col_values(2))-1])
    list_oil_log.append(table.col_values(3)[len(table.col_values(3))-1])
    list_oil_log.append(table.col_values(0)[len(table.col_values(0))-1])
    return list_oil_log
def time_loacl():
    local_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    return local_time
def input_comm():
        input_comm = "请输入命令:"
        print(input_comm)
        command = input()
        return int(command)
# -----------------------------------main-------------------------------------------------
if __name__=='__main__':
    execl_oil=execl_11()
    intrduce()
    while 1:
        input_comm = "请输入命令:"
        print(input_comm)
        command = int(input())
        if command == 0:
            intrduce()
        elif command == 1:
            print("【提示】请关闭 oil_log.xlsx , 如果已经关闭，请忽略")
            time.sleep(1)
            print("是否更新execl表格,【是】回车键 【否】任意键加回车键")
            command_value = input("")
            if command_value:
                print("请输入1号油罐的高度(m):")
                oil_H1 = float(input(""))
                oil_current1 = (calu_oil_L(oil_H1))
                print("请输入2号油罐的高度(m):")
                oil_H2 = float(input(""))
                oil_current2 = (calu_oil_L(oil_H2))
                print("oil_H1:",oil_current1)
                print("oil_H2:",oil_current2)
            else:
                print("请输入1号油罐的高度(m):")
                oil_H1 = float(input(""))
                oil_current1 = (calu_oil_L(oil_H1))
                print("请输入2号油罐的高度(m):")
                oil_H2 = float(input(""))
                oil_current2 = (calu_oil_L(oil_H2))
                print("1号油罐:",oil_current1)
                print("2号油罐:",oil_current2)
                update_oil1 = oil_current1
                update_oil2 = oil_current2
                oil_H_add()
                print("更新成功!")
                intrduce()

        elif command == 2:
            pass
        elif command == 3:
            quit()
        else:
            print("---命令不对---")